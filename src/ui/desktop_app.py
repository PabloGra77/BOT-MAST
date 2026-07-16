# -*- coding: utf-8 -*-
"""
BOT - Aplicacion Desktop Principal
Panel de control con gestion de credenciales y navegadores
"""
from __future__ import annotations
import sys
import json
import os
import threading
import subprocess
from pathlib import Path
from typing import Optional, List

# ---------------------------------------------------------------------------
# Asegurar que src/ este en el path
# ---------------------------------------------------------------------------
_SRC_DIR = Path(__file__).resolve().parent.parent   # src/
_ROOT_DIR = _SRC_DIR.parent                          # BOT/
if str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))
if str(_ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(_ROOT_DIR))

# Cuando corre como .exe (PyInstaller), __file__ apunta a la carpeta temporal
# _MEIxxxx. config/, logs/, downloads/ deben vivir junto al .exe para persistir.
_BASE_OVERRIDE = os.environ.get("BOT_BASE_DIR")
if _BASE_OVERRIDE:
    _DATA_DIR = Path(_BASE_OVERRIDE).resolve()
elif getattr(sys, "frozen", False):
    _DATA_DIR = Path(sys.executable).resolve().parent
else:
    _DATA_DIR = _ROOT_DIR

try:
    from PyQt6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QTabWidget, QPushButton, QLabel, QLineEdit, QComboBox, QSpinBox,
        QTextEdit, QTableWidget, QTableWidgetItem, QMessageBox, QDialog,
        QFormLayout, QDialogButtonBox, QCheckBox, QGroupBox, QFrame,
        QSizePolicy, QSlider, QHeaderView, QAbstractItemView,
        QListWidget, QListWidgetItem, QStatusBar, QFileDialog, QProgressBar,
        QScrollArea
    )
    from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QDateTime, QSize
    from PyQt6.QtGui import QFont, QTextCursor
except ImportError as _e_pyqt:
    print(f"ERROR: No se pudo cargar PyQt6: {_e_pyqt!r}")
    import traceback as _tb_pyqt
    _tb_pyqt.print_exc()
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config helpers
# ---------------------------------------------------------------------------
_CONFIG_PATH = _DATA_DIR / "config" / "config.json"

def _load_config() -> dict:
    if _CONFIG_PATH.exists():
        with open(_CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def _save_config(cfg: dict):
    _CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=4, ensure_ascii=False)

try:
    from bot_app.common.settings import CONFIG_PATH as _SP
    def load_main_config() -> dict:
        return _load_config()
    def save_main_config(cfg: dict):
        _save_config(cfg)
except ImportError:
    load_main_config = _load_config
    save_main_config = _save_config

VERSION = "1.0.4"
# Repositorio para buscar/descargar nuevas versiones (auto-actualizacion).
GITHUB_REPO = "PabloGra77/BOT-MAST"
# Maximo de navegadores/usuarios en paralelo. Cada navegador requiere un
# usuario INPEC distinto (credenciales_hc) para ser estable.
MAX_BROWSERS = 8

# ===========================================================================
# Estilos
# ===========================================================================
DARK_STYLE = """
* { font-family: "Segoe UI", "Inter", Arial, sans-serif; }
QMainWindow, QDialog {background-color: #0d1117; color: #e6edf3;}
QWidget {background-color: #0d1117; color: #e6edf3; font-size: 13px;}
QToolTip {background: #161b22; color: #e6edf3; border: 1px solid #30363d; padding: 4px;}

/* Pestañas */
QTabWidget::pane {border: 1px solid #30363d; background: #0d1117; border-radius: 8px; top: -1px;}
QTabBar::tab {background: transparent; color: #8b949e; padding: 9px 20px; margin-right: 2px;
    border: none; border-bottom: 2px solid transparent; font-weight: 600;}
QTabBar::tab:selected {color: #e6edf3; border-bottom: 2px solid #2f81f7;}
QTabBar::tab:hover {color: #e6edf3;}

/* Botones: por defecto neutro; .primary azul; .success verde; .danger rojo */
QPushButton {
    background: #21262d; color: #e6edf3; border: 1px solid #30363d;
    border-radius: 6px; padding: 8px 16px; font-weight: 600;}
QPushButton:hover {background: #30363d; border-color: #3d444d;}
QPushButton:pressed {background: #161b22;}
QPushButton:disabled {background: #161b22; color: #6e7681; border-color: #21262d;}
QPushButton[class="primary"] {background: #2f81f7; color: #ffffff; border: 1px solid #2f81f7;}
QPushButton[class="primary"]:hover {background: #388bfd; border-color: #388bfd;}
QPushButton[class="primary"]:pressed {background: #1f6feb;}
QPushButton[class="success"] {background: #238636; color: #ffffff; border: 1px solid #2ea043;}
QPushButton[class="success"]:hover {background: #2ea043;}
QPushButton[class="danger"] {background: #21262d; color: #f85149; border: 1px solid #da3633;}
QPushButton[class="danger"]:hover {background: #da3633; color: #ffffff;}

/* Campos */
QLineEdit, QTextEdit, QComboBox, QSpinBox {
    background: #0d1117; color: #e6edf3; border: 1px solid #30363d;
    border-radius: 6px; padding: 6px 8px; selection-background-color: #1f6feb;}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QSpinBox:focus {border: 1px solid #2f81f7;}
QComboBox::drop-down {border: none; width: 22px;}
QComboBox QAbstractItemView {background: #161b22; color: #e6edf3; border: 1px solid #30363d;
    selection-background-color: #2f81f7;}

/* Grupos / tarjetas */
QGroupBox {background: #161b22; border: 1px solid #30363d; border-radius: 10px;
    margin-top: 14px; padding: 14px 12px 10px 12px; font-weight: 700; color: #e6edf3;}
QGroupBox::title {subcontrol-origin: margin; left: 12px; top: 2px; padding: 0 6px; color: #8b949e;}

/* Tablas */
QTableWidget {background: #161b22; color: #e6edf3; gridline-color: #21262d;
    border: 1px solid #30363d; border-radius: 8px;
    selection-background-color: #1f6feb; selection-color: #ffffff;}
QTableWidget::item {padding: 4px;}
QHeaderView::section {background: #21262d; color: #8b949e; padding: 8px; border: none;
    border-bottom: 1px solid #30363d; font-weight: 600;}

/* Scrollbars */
QScrollBar:vertical {background: transparent; width: 12px; margin: 2px;}
QScrollBar::handle:vertical {background: #30363d; border-radius: 6px; min-height: 30px;}
QScrollBar::handle:vertical:hover {background: #3d444d;}
QScrollBar:horizontal {background: transparent; height: 12px; margin: 2px;}
QScrollBar::handle:horizontal {background: #30363d; border-radius: 6px; min-width: 30px;}
QScrollBar::add-line, QScrollBar::sub-line {height: 0; width: 0;}
QScrollBar::add-page, QScrollBar::sub-page {background: transparent;}
QScrollArea {border: none; background: transparent;}

/* Checks / sliders */
QCheckBox {spacing: 8px;}
QCheckBox::indicator {width: 18px; height: 18px; border: 1px solid #30363d; border-radius: 4px; background: #0d1117;}
QCheckBox::indicator:checked {background: #2f81f7; border-color: #2f81f7;}
QSlider::groove:horizontal {background: #21262d; height: 6px; border-radius: 3px;}
QSlider::handle:horizontal {background: #2f81f7; width: 16px; height: 16px;
    margin: -6px 0; border-radius: 8px;}
QSlider::sub-page:horizontal {background: #2f81f7; border-radius: 3px;}

/* Barra de estado / progreso */
QStatusBar {background: #161b22; color: #8b949e; border-top: 1px solid #30363d;}
QProgressBar {background: #21262d; border: 1px solid #30363d; border-radius: 6px;
    text-align: center; color: #e6edf3; height: 18px;}
QProgressBar::chunk {background: #2f81f7; border-radius: 5px;}
"""


# ===========================================================================
# Tab: Credenciales
# ===========================================================================
# Valores fijos (siempre los mismos para INPEC)
INPEC_URL_FIJA   = "https://sisipec.salud360.app/Inpec360/servlet/ingreso"
INPEC_VERIF_FIJO = "Inpec"


class CredencialesTab(QWidget):
    """
    Tabla unificada: cada fila = un usuario = un navegador.
    El usuario #1 actúa como Principal (también para Agendas).
    Verificación y URL se asignan automáticamente (siempre 'Inpec' y la URL fija).
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self._users: List[dict] = []
        self._build_ui()
        self.load_from_config()

    def _build_ui(self):
        main = QVBoxLayout(self)

        titulo = QLabel("Usuarios INPEC")
        titulo.setStyleSheet("color:#2f81f7; font-size:16px; font-weight:bold;")
        main.addWidget(titulo)

        info = QLabel(
            "Cada usuario que agregues abre un navegador para descargar en paralelo.\n"
            "El primer usuario también se usa para crear Agendas.\n"
            f"Verificación fija: <b>{INPEC_VERIF_FIJO}</b>  ·  "
            f"URL fija: <code>{INPEC_URL_FIJA}</code>\n"
            "Los datos se guardan automáticamente en config.json y persisten al cerrar."
        )
        info.setWordWrap(True)
        info.setStyleSheet("color:#8b949e; font-size:11px; padding:4px;")
        main.addWidget(info)

        self.lbl_count = QLabel("Sin usuarios guardados.")
        self.lbl_count.setStyleSheet(
            "color:#e6edf3; font-size:13px; font-weight:bold; padding:6px; "
            "background:#0d1117; border:1px solid #30363d; border-radius:6px;"
        )
        main.addWidget(self.lbl_count)

        self.tbl = QTableWidget(0, 4)
        self.tbl.setHorizontalHeaderLabels(["#", "Rol", "Usuario", "Contraseña"])
        self.tbl.setColumnWidth(0, 50)
        self.tbl.setColumnWidth(1, 130)
        self.tbl.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.tbl.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl.setMinimumHeight(220)
        self.tbl.doubleClicked.connect(lambda *_: self._edit())
        main.addWidget(self.tbl)

        rb = QHBoxLayout()
        for lbl_btn, slot, cls in [
            ("➕ Agregar Usuario", self._add,  "success"),
            ("✏ Editar",          self._edit, ""),
            ("🗑 Eliminar",        self._del,  "danger"),
            ("🌐 Probar URL",      self._probar_url, ""),
        ]:
            b = QPushButton(lbl_btn)
            if cls:
                b.setProperty("class", cls)
            b.clicked.connect(slot)
            rb.addWidget(b)
        main.addLayout(rb)

        self.lbl_st = QLabel("")
        self.lbl_st.setStyleSheet("color:#3fb950; font-size:11px;")
        main.addWidget(self.lbl_st)
        main.addStretch()

    # ---------- helpers ----------
    def _refresh_table(self):
        self.tbl.setRowCount(0)
        for r, c in enumerate(self._users):
            self.tbl.insertRow(r)
            self.tbl.setItem(r, 0, QTableWidgetItem(f"Nav. {r+1}"))
            rol = "⭐ Principal + HC" if r == 0 else "HC (paralelo)"
            self.tbl.setItem(r, 1, QTableWidgetItem(rol))
            self.tbl.setItem(r, 2, QTableWidgetItem(c.get("usuario", "")))
            masked = "•" * max(len(c.get("contrasena", "")), 4)
            self.tbl.setItem(r, 3, QTableWidgetItem(masked))
        n = len(self._users)
        if n == 0:
            self.lbl_count.setText("⚠ Sin usuarios guardados — pulsa 'Agregar Usuario' para empezar.")
            self.lbl_count.setStyleSheet(
                "color:#f59e0b; font-size:13px; font-weight:bold; padding:6px; "
                "background:#0d1117; border:1px solid #f59e0b; border-radius:6px;"
            )
        elif n == 1:
            self.lbl_count.setText(
                f"✅ 1 usuario guardado  →  1 navegador  (sin paralelismo HC)"
            )
            self.lbl_count.setStyleSheet(
                "color:#22c55e; font-size:13px; font-weight:bold; padding:6px; "
                "background:#0d1117; border:1px solid #22c55e; border-radius:6px;"
            )
        else:
            self.lbl_count.setText(
                f"✅ {n} usuarios guardados  →  {n} navegadores en paralelo"
            )
            self.lbl_count.setStyleSheet(
                "color:#22c55e; font-size:13px; font-weight:bold; padding:6px; "
                "background:#0d1117; border:1px solid #22c55e; border-radius:6px;"
            )

    def _status(self, msg, color="#3fb950"):
        self.lbl_st.setText(msg)
        self.lbl_st.setStyleSheet(f"color:{color}; font-size:11px;")

    def _persistir(self):
        """Guarda inmediatamente en config.json (no requiere botón aparte)."""
        cfg = load_main_config()
        if self._users:
            principal = self._users[0]
            cfg["credenciales"] = {
                "usuario":             principal.get("usuario", ""),
                "contrasena":          principal.get("contrasena", ""),
                "numero_verificacion": INPEC_VERIF_FIJO,
                "url":                 INPEC_URL_FIJA,
            }
            cfg["credenciales_hc"] = [
                {
                    "usuario":             u.get("usuario", ""),
                    "contrasena":          u.get("contrasena", ""),
                    "numero_verificacion": INPEC_VERIF_FIJO,
                    "url":                 INPEC_URL_FIJA,
                }
                for u in self._users
            ]
        else:
            cfg["credenciales"] = {}
            cfg["credenciales_hc"] = []
        save_main_config(cfg)

    # ---------- carga / persistencia ----------
    def load_from_config(self):
        cfg = load_main_config()
        usuarios = []
        # 1. Principal (si existe)
        principal = cfg.get("credenciales", {}) or {}
        if principal.get("usuario"):
            usuarios.append({
                "usuario":    str(principal.get("usuario", "")).strip().upper(),
                "contrasena": principal.get("contrasena", ""),
            })
        # 2. HC adicionales (evitar duplicar el principal)
        for h in cfg.get("credenciales_hc", []) or []:
            u = str(h.get("usuario", "")).strip().upper()
            if u and not any(x["usuario"] == u for x in usuarios):
                usuarios.append({
                    "usuario":    u,
                    "contrasena": h.get("contrasena", ""),
                })
        self._users = usuarios
        self._refresh_table()

    # ---------- acciones ----------
    def _add(self):
        if len(self._users) >= MAX_BROWSERS:
            QMessageBox.warning(self, "Límite",
                f"Máximo {MAX_BROWSERS} usuarios (= {MAX_BROWSERS} navegadores en paralelo).")
            return
        dlg = _CredDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get_data()
            if not d["usuario"]:
                return
            if any(u["usuario"] == d["usuario"] for u in self._users):
                QMessageBox.warning(self, "Duplicado",
                    f"Ya existe un usuario '{d['usuario']}'.")
                return
            self._users.append(d)
            self._refresh_table()
            self._persistir()
            self._status(
                f"➕ Agregado '{d['usuario']}'  ·  guardado en config.json  "
                f"·  total: {len(self._users)} navegador(es)")

    def _edit(self):
        r = self.tbl.currentRow()
        if r < 0:
            QMessageBox.information(self, "Selecciona",
                "Selecciona primero una fila (o haz doble clic).")
            return
        dlg = _CredDialog(self, self._users[r])
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._users[r] = dlg.get_data()
            self._refresh_table()
            self._persistir()
            self._status(
                f"✏ Modificado '{self._users[r]['usuario']}'  ·  guardado en config.json")

    def _del(self):
        r = self.tbl.currentRow()
        if r < 0:
            QMessageBox.information(self, "Selecciona", "Selecciona primero una fila.")
            return
        u = self._users[r].get("usuario", "")
        msg = f"¿Eliminar el usuario '{u}'?"
        if r == 0 and len(self._users) > 1:
            msg += "\n\nEs el usuario Principal — el siguiente pasará a serlo."
        if QMessageBox.question(self, "Eliminar", msg) == QMessageBox.StandardButton.Yes:
            self._users.pop(r)
            self._refresh_table()
            self._persistir()
            self._status(f"🗑 Eliminado '{u}'  ·  guardado en config.json", "#f59e0b")

    def _probar_url(self):
        import webbrowser
        webbrowser.open(INPEC_URL_FIJA)

    # ---------- compatibilidad con MainWindow._save_all ----------
    def save_main_cred(self):
        self._persistir()

    def _save_hc(self):
        self._persistir()


class _CredDialog(QDialog):
    def __init__(self, parent=None, data=None):
        super().__init__(parent)
        self.setWindowTitle("Usuario INPEC — Navegador")
        self.setMinimumWidth(440)
        self.setStyleSheet(DARK_STYLE)
        d = data or {}
        lay = QVBoxLayout(self)

        info = QLabel(
            "Cada usuario corresponde a un navegador para descargas en paralelo.\n"
            f"• Verificación: <b>{INPEC_VERIF_FIJO}</b>  (automática)\n"
            f"• URL: <code>{INPEC_URL_FIJA}</code>  (automática)"
        )
        info.setStyleSheet("color:#8b949e; font-size:11px; padding:4px;")
        info.setWordWrap(True)
        lay.addWidget(info)

        form = QFormLayout()
        self.txt_u = QLineEdit(d.get("usuario", ""))
        self.txt_u.setPlaceholderText("Ej: PGRANADOS")
        self.txt_p = QLineEdit(d.get("contrasena", ""))
        self.txt_p.setEchoMode(QLineEdit.EchoMode.Password)
        self.txt_p.setPlaceholderText("Contraseña del usuario")

        rp = QHBoxLayout()
        rp.addWidget(self.txt_p)
        btn_show = QPushButton("Mostrar")
        btn_show.setFixedWidth(72)
        def _toggle():
            if self.txt_p.echoMode() == QLineEdit.EchoMode.Password:
                self.txt_p.setEchoMode(QLineEdit.EchoMode.Normal); btn_show.setText("Ocultar")
            else:
                self.txt_p.setEchoMode(QLineEdit.EchoMode.Password); btn_show.setText("Mostrar")
        btn_show.clicked.connect(_toggle)
        rp.addWidget(btn_show)
        wp = QWidget(); wp.setLayout(rp)

        form.addRow("Usuario:", self.txt_u)
        form.addRow("Contraseña:", wp)
        lay.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Guardar")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Cancelar")
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def _on_accept(self):
        if not self.txt_u.text().strip():
            QMessageBox.warning(self, "Falta usuario", "El usuario no puede estar vacío.")
            return
        if not self.txt_p.text():
            QMessageBox.warning(self, "Falta contraseña", "La contraseña no puede estar vacía.")
            return
        self.accept()

    def get_data(self):
        return {
            "usuario":    self.txt_u.text().strip().upper(),
            "contrasena": self.txt_p.text(),
        }


# ===========================================================================
# Tab: Navegadores
# ===========================================================================
class NavegadoresTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self.load_from_config()

    def _build_ui(self):
        main = QVBoxLayout(self)

        grp_tipo = QGroupBox("Tipo de Navegador")
        main.addWidget(grp_tipo)
        fl = QFormLayout(grp_tipo)
        self.cmb = QComboBox()
        self.cmb.addItems(["chrome", "edge"])
        self.cmb.setToolTip("Selecciona el navegador instalado en este equipo")
        fl.addRow("Navegador:", self.cmb)

        grp_cant = QGroupBox(f"Cantidad de Navegadores Paralelos  (max {MAX_BROWSERS})")
        main.addWidget(grp_cant)
        vl2 = QVBoxLayout(grp_cant)

        lbl_info = QLabel(
            "Cada navegador ejecuta una sesion independiente.\n"
            "Mas navegadores = mayor paralelismo pero mas uso de CPU/RAM."
        )
        lbl_info.setStyleSheet("color:#8b949e; font-size:11px;")
        vl2.addWidget(lbl_info)

        row_sl = QHBoxLayout()
        self.sld = QSlider(Qt.Orientation.Horizontal)
        self.sld.setMinimum(1)
        self.sld.setMaximum(MAX_BROWSERS)
        self.sld.setValue(1)
        self.sld.setTickPosition(QSlider.TickPosition.TicksBelow)
        self.sld.setTickInterval(1)
        self.sld.valueChanged.connect(self._on_slide)
        self.lbl_n = QLabel("1")
        self.lbl_n.setStyleSheet("color:#2f81f7; font-size:26px; font-weight:bold;")
        self.lbl_n.setFixedWidth(36)
        self.lbl_n.setAlignment(Qt.AlignmentFlag.AlignCenter)
        row_sl.addWidget(self.sld)
        row_sl.addWidget(self.lbl_n)
        vl2.addLayout(row_sl)

        # iconos visuales
        self.icons_row = QHBoxLayout()
        self._icons: List[QLabel] = []
        for _ in range(MAX_BROWSERS):
            ico = QLabel("O")
            ico.setFixedSize(QSize(44, 44))
            ico.setAlignment(Qt.AlignmentFlag.AlignCenter)
            ico.setStyleSheet("font-size:22px; background:#161b22; border-radius:8px; border:1px solid #30363d;")
            self._icons.append(ico)
            self.icons_row.addWidget(ico)
        self.icons_row.addStretch()
        vl2.addLayout(self.icons_row)

        self.lbl_req = QLabel("")
        self.lbl_req.setStyleSheet("color:#8b949e; font-size:11px; font-style:italic;")
        self.lbl_req.setWordWrap(True)
        vl2.addWidget(self.lbl_req)

        grp_adv = QGroupBox("Carpeta de Descargas")
        main.addWidget(grp_adv)
        fl2 = QFormLayout(grp_adv)

        self.txt_dl = QLineEdit()
        self.txt_dl.setReadOnly(True)
        self.txt_dl.setPlaceholderText("Dejar vacio para usar carpeta predeterminada")
        btn_br = QPushButton("...")
        btn_br.setFixedWidth(32)
        btn_br.clicked.connect(self._browse)
        row_dl = QHBoxLayout()
        row_dl.addWidget(self.txt_dl)
        row_dl.addWidget(btn_br)
        w_dl = QWidget(); w_dl.setLayout(row_dl)
        fl2.addRow("Carpeta descargas:", w_dl)

        grp_check = QGroupBox("Estado de Navegadores Instalados")
        main.addWidget(grp_check)
        vl3 = QVBoxLayout(grp_check)
        r_st = QHBoxLayout()
        self.lbl_ch = QLabel("Chrome: --")
        self.lbl_ed = QLabel("Edge: --")
        r_st.addWidget(self.lbl_ch)
        r_st.addWidget(self.lbl_ed)
        vl3.addLayout(r_st)
        btn_verify = QPushButton("Verificar Navegadores Instalados")
        btn_verify.clicked.connect(self.check_browsers)
        vl3.addWidget(btn_verify)

        # === Escaneo de capacidad del PC =====================================
        grp_cap = QGroupBox("Capacidad del Equipo")
        main.addWidget(grp_cap)
        vl_cap = QVBoxLayout(grp_cap)
        self.lbl_cap_info = QLabel(
            "Pulsa el boton para escanear CPU, RAM y disco y obtener "
            "una recomendacion del numero de navegadores en paralelo."
        )
        self.lbl_cap_info.setStyleSheet("color:#8b949e; font-size:11px;")
        self.lbl_cap_info.setWordWrap(True)
        vl_cap.addWidget(self.lbl_cap_info)

        self.lbl_cap_result = QLabel("")
        self.lbl_cap_result.setStyleSheet(
            "font-size:12px; padding:6px; background:#161b22; "
            "border:1px solid #30363d; border-radius:6px;"
        )
        self.lbl_cap_result.setWordWrap(True)
        self.lbl_cap_result.setVisible(False)
        vl_cap.addWidget(self.lbl_cap_result)

        row_cap = QHBoxLayout()
        btn_scan = QPushButton("Escanear Capacidad del PC")
        btn_scan.setProperty("class", "primary")
        btn_scan.clicked.connect(self.escanear_capacidad)
        row_cap.addWidget(btn_scan)
        self.btn_apply_cap = QPushButton("Aplicar Recomendacion")
        self.btn_apply_cap.setProperty("class", "success")
        self.btn_apply_cap.setEnabled(False)
        self.btn_apply_cap.clicked.connect(self._aplicar_recomendacion)
        row_cap.addWidget(self.btn_apply_cap)
        vl_cap.addLayout(row_cap)
        self._recomendado = None

        # Boton para liberar RAM cerrando apps del usuario no esenciales.
        row_force = QHBoxLayout()
        self.btn_forzar = QPushButton("Forzar Equipo (cerrar apps y liberar RAM)")
        self.btn_forzar.setProperty("class", "danger")
        self.btn_forzar.setToolTip(
            "Cierra las aplicaciones abiertas del usuario (navegadores, Office, etc.)\n"
            "para liberar RAM y dedicar el equipo al bot.\n"
            "NO toca procesos del sistema de Windows ni el propio bot.\n"
            "Guarda tu trabajo antes de usarlo."
        )
        self.btn_forzar.clicked.connect(self._forzar_equipo)
        row_force.addWidget(self.btn_forzar)
        vl_cap.addLayout(row_force)
        # =====================================================================

        row_save = QHBoxLayout()
        row_save.addStretch()
        btn_save = QPushButton("Guardar Configuracion de Navegadores")
        btn_save.setProperty("class", "success")
        btn_save.clicked.connect(self.save_config)
        row_save.addWidget(btn_save)
        main.addLayout(row_save)

        self.lbl_st = QLabel("")
        self.lbl_st.setStyleSheet("color:#3fb950; font-size:11px;")
        main.addWidget(self.lbl_st)
        main.addStretch()

        # init icons
        self._on_slide(1)

    def _on_slide(self, v):
        self.lbl_n.setText(str(v))
        for i, ico in enumerate(self._icons):
            if i < v:
                ico.setStyleSheet(
                    "font-size:22px; background:#30363d; border-radius:8px; "
                    "border:2px solid #2f81f7; color:#e6edf3;"
                )
            else:
                ico.setStyleSheet(
                    "font-size:22px; background:#161b22; border-radius:8px; "
                    "border:1px solid #30363d; color:#444;"
                )
        if v == 1:
            self.lbl_req.setText("Con 1 navegador solo se necesita la credencial principal.")
        else:
            self.lbl_req.setText(
                f"Con {v} navegadores se usan credenciales HC ({v} credenciales en total). "
                "Configuralas en la pestana Credenciales."
            )

    def _browse(self):
        d = QFileDialog.getExistingDirectory(self, "Carpeta de descargas")
        if d:
            self.txt_dl.setText(d)

    def load_from_config(self):
        cfg = load_main_config()
        nav = cfg.get("navegadores", {})
        idx = self.cmb.findText(nav.get("tipo", "chrome"))
        self.cmb.setCurrentIndex(max(0, idx))
        cantidad = max(1, min(MAX_BROWSERS, int(nav.get("cantidad_max", 1))))
        self.sld.setValue(cantidad)
        self.txt_dl.setText(nav.get("carpeta_descargas", ""))
        self._on_slide(cantidad)

    def save_config(self):
        cfg = load_main_config()
        cfg["navegadores"] = {
            "tipo":              self.cmb.currentText(),
            "cantidad_max":      self.sld.value(),
            "carpeta_descargas": self.txt_dl.text().strip(),
        }
        save_main_config(cfg)
        self.lbl_st.setText(
            f"Guardado: {self.sld.value()} navegador(es) [{self.cmb.currentText()}]"
        )

    def check_browsers(self):
        import shutil
        def _ok(paths):
            return any(Path(p).exists() for p in paths)

        ch = _ok([
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ])
        ed = _ok([
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        ])
        self.lbl_ch.setText(f"Chrome: {'Instalado' if ch else 'No encontrado'}")
        self.lbl_ed.setText(f"Edge: {'Instalado' if ed else 'No encontrado'}")
        if not ch and not ed:
            QMessageBox.warning(
                self, "Sin navegadores",
                "No se detecto Chrome ni Edge en este equipo.\n"
                "Instala al menos uno para usar el bot.\n\n"
                "Chrome: https://www.google.com/chrome/\n"
                "Edge: https://www.microsoft.com/es-es/edge"
            )

    # ------------------------------------------------------------------
    # Escaneo de capacidad del PC
    # ------------------------------------------------------------------
    def escanear_capacidad(self):
        """Inspecciona CPU, RAM y disco para recomendar # de navegadores."""
        try:
            import psutil
        except ImportError:
            QMessageBox.critical(
                self, "psutil no instalado",
                "Falta el modulo 'psutil'. Reinstala el ejecutable o ejecuta:\n"
                "pip install psutil"
            )
            return

        try:
            cpu_logicos = psutil.cpu_count(logical=True) or 1
            cpu_fisicos = psutil.cpu_count(logical=False) or cpu_logicos
            mem = psutil.virtual_memory()
            ram_total_gb = mem.total / (1024 ** 3)
            ram_libre_gb = mem.available / (1024 ** 3)
            cpu_uso = psutil.cpu_percent(interval=0.6)
            disco = psutil.disk_usage(os.path.expanduser("~"))
            disco_libre_gb = disco.free / (1024 ** 3)
        except Exception as exc:
            QMessageBox.critical(self, "Error de escaneo", f"No se pudo leer info del sistema:\n{exc}")
            return

        # Reglas de recomendacion. Un navegador (1 pestaña automatizando INPEC)
        # consume ~0.7-0.9 GB de RAM. El factor que casi siempre manda es la RAM
        # LIBRE en el momento del escaneo (no la CPU).
        ram_para_bot_gb = max(0.5, ram_libre_gb - 1.0)   # dejar 1.0 GB libre de colchon
        cap_por_ram = int(ram_para_bot_gb // 0.85)       # ~0.85 GB por navegador
        cap_por_cpu = max(1, cpu_logicos // 2)           # 2 hilos logicos por nav
        # La CPU solo penaliza si esta GENUINAMENTE ocupada (>60%). Una carga
        # momentanea baja no debe limitar un equipo con muchos nucleos.
        if cpu_uso >= 60:
            cap_por_load = max(1, int((100 - cpu_uso) / 15))
        else:
            cap_por_load = cap_por_cpu
        recomendado = max(1, min(MAX_BROWSERS, cap_por_ram, cap_por_cpu, cap_por_load))

        # Determinar el factor limitante para mostrarlo (transparencia).
        _factores = {"RAM disponible": cap_por_ram, "nucleos de CPU": cap_por_cpu, "carga de CPU": cap_por_load}
        factor_limite = min(_factores, key=_factores.get)

        if disco_libre_gb < 2:
            recomendado = 1
            warn_disco = "  (espacio en disco bajo: solo 1 navegador)"
        else:
            warn_disco = ""

        self._recomendado = recomendado
        self.btn_apply_cap.setEnabled(True)

        # Categoria del equipo
        if recomendado >= 4:
            categoria = "ALTA"
            color = "#3fb950"
        elif recomendado >= 2:
            categoria = "MEDIA"
            color = "#d29922"
        else:
            categoria = "BASICA"
            color = "#2f81f7"

        html = (
            f"<div style='color:{color}; font-weight:bold; font-size:13px;'>"
            f"&#9989; Recomendado: <span style='font-size:18px;'>{recomendado}</span> "
            f"navegador{'es' if recomendado>1 else ''} en paralelo "
            f"&middot; Capacidad: {categoria}</div>"
            f"<br>"
            f"<b>CPU:</b> {cpu_logicos} nucleos logicos ({cpu_fisicos} fisicos) "
            f"&middot; uso actual: {cpu_uso:.0f}%<br>"
            f"<b>RAM:</b> {ram_libre_gb:.1f} GB libres / {ram_total_gb:.1f} GB totales<br>"
            f"<b>Disco (perfil):</b> {disco_libre_gb:.1f} GB libres{warn_disco}<br>"
            f"<b>Factor limitante:</b> {factor_limite}<br>"
            f"<br>"
            + (
                f"<i>El limite es la <b>RAM libre</b> ({ram_libre_gb:.1f} GB). "
                f"Cierra otros programas y vuelve a escanear para permitir mas "
                f"navegadores (cada uno usa ~0.85 GB).</i>"
                if factor_limite == "RAM disponible" else
                f"<i>Cada navegador usa ~0.85 GB RAM y ~2 hilos logicos. "
                f"Si el PC se siente lento, reduce a {max(1, recomendado-1)}.</i>"
            )
        )
        self.lbl_cap_result.setText(html)
        self.lbl_cap_result.setVisible(True)

    def _aplicar_recomendacion(self):
        if not self._recomendado:
            return
        v = max(1, min(MAX_BROWSERS, int(self._recomendado)))
        self.sld.setValue(v)
        self._on_slide(v)
        self.lbl_st.setText(f"Aplicada recomendacion: {v} navegador(es). Recuerda Guardar.")

    # Lista BLANCA de aplicaciones pesadas del usuario que SÍ se cierran al
    # "Forzar Equipo" (en minusculas, con .exe). Son los grandes consumidores de
    # RAM. NO se tocan drivers (audio/touchpad/GPU), seguridad/VPN, OneDrive,
    # agentes de trabajo (UiPath) ni procesos del sistema.
    _PROC_CERRAR = {
        # Navegadores y motores web embebidos
        "chrome.exe", "msedge.exe", "msedgewebview2.exe", "firefox.exe",
        "opera.exe", "opera_gx.exe", "brave.exe", "vivaldi.exe", "iexplore.exe",
        "chromium.exe", "browser.exe",
        # Microsoft Office / correo
        "winword.exe", "excel.exe", "powerpnt.exe", "outlook.exe", "olk.exe",
        "onenote.exe", "msaccess.exe", "mspub.exe", "visio.exe", "winproj.exe",
        "m365copilot.exe",
        # Comunicaciones
        "teams.exe", "ms-teams.exe", "slack.exe", "discord.exe", "zoom.exe",
        "skype.exe", "telegram.exe", "whatsapp.exe", "webex.exe", "webexmta.exe",
        # Multimedia
        "spotify.exe", "spotifylauncher.exe", "vlc.exe", "itunes.exe",
        "wmplayer.exe", "groove.exe",
        # IA / notas / PDF / creatividad
        "chatgpt.exe", "claude.exe", "notion.exe", "obsidian.exe", "evernote.exe",
        "acrobat.exe", "acrord32.exe", "photoshop.exe", "illustrator.exe",
        # Juegos / tiendas / nube (grandes consumidores)
        "steam.exe", "steamwebhelper.exe", "epicgameslauncher.exe",
        "battle.net.exe", "dropbox.exe",
        # Utilidades varias
        "ccleaner.exe",
    }

    def _forzar_equipo(self):
        """Cierra apps PESADAS del usuario (lista blanca) para liberar RAM y
        dedicar el equipo al bot. Salvaguardas: solo cierra apps de la lista
        blanca, solo del usuario actual (nunca SYSTEM/SERVICE), nunca el propio
        proceso ni sus ancestros. NO toca drivers, seguridad, VPN ni OneDrive."""
        try:
            import psutil
        except ImportError:
            QMessageBox.critical(self, "psutil no instalado", "Falta el modulo 'psutil'.")
            return

        resp = QMessageBox.warning(
            self, "Forzar Equipo",
            "Esto CERRARA las aplicaciones pesadas abiertas (navegadores, Office, "
            "Teams/Slack, Spotify, etc.) para liberar RAM y dedicar el equipo al bot.\n\n"
            "• NO cierra Windows, drivers, antivirus/VPN, OneDrive ni el propio bot.\n"
            "• GUARDA tu trabajo antes: lo no guardado se PIERDE.\n"
            "• Hazlo ANTES de iniciar el bot (tambien cierra navegadores).\n\n"
            "¿Continuar?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if resp != QMessageBox.StandardButton.Yes:
            return

        # Usuario actual y arbol propio de procesos (no cerrarse a si mismo).
        protegidos_pids = set()
        usuario_actual = ""
        try:
            yo = psutil.Process(os.getpid())
            usuario_actual = (yo.username() or "").lower()
            p = yo
            for _ in range(12):  # tope defensivo
                if p is None:
                    break
                protegidos_pids.add(p.pid)
                try:
                    p = p.parent()
                except Exception:
                    break
        except Exception:
            pass

        mem_antes = psutil.virtual_memory().available / (1024 ** 3)

        objetivos = []  # [(proc, nombre)]
        for proc in psutil.process_iter(["pid", "name", "username"]):
            try:
                pid = proc.info.get("pid")
                nombre = (proc.info.get("name") or "")
                nl = nombre.lower()
                if not pid or pid in protegidos_pids:
                    continue
                # Solo apps de la lista blanca.
                if nl not in self._PROC_CERRAR:
                    continue
                uname = (proc.info.get("username") or "").lower()
                # CLAVE: solo procesos del usuario actual (jamas SYSTEM/SERVICE).
                if not usuario_actual or uname != usuario_actual:
                    continue
                objetivos.append((proc, nombre))
            except Exception:
                continue

        # Cierre suave (terminate) y luego forzado (kill) de lo que quede.
        cerrados = []
        for proc, nombre in objetivos:
            try:
                proc.terminate()
                cerrados.append(nombre)
            except Exception:
                pass
        try:
            _gone, vivos = psutil.wait_procs([p for p, _ in objetivos], timeout=3)
            for proc in vivos:
                try:
                    proc.kill()
                except Exception:
                    pass
        except Exception:
            pass

        mem_despues = psutil.virtual_memory().available / (1024 ** 3)
        from collections import Counter
        conteo = Counter(c.lower() for c in cerrados)
        detalle = ", ".join(
            (f"{n} (x{c})" if c > 1 else n) for n, c in sorted(conteo.items())
        ) or "ninguna"

        QMessageBox.information(
            self, "Forzar Equipo",
            f"Aplicaciones cerradas: {len(cerrados)}\n"
            f"RAM libre: {mem_antes:.1f} GB  ->  {mem_despues:.1f} GB\n\n"
            f"Cerradas: {detalle}\n\n"
            f"Se vuelve a escanear la capacidad automaticamente."
        )
        try:
            self.escanear_capacidad()
        except Exception:
            pass


# ===========================================================================
# Tab: Dashboard
# ===========================================================================
class DashboardTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        hdr = QLabel("BOT  |  Panel de Control")
        hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hdr.setStyleSheet("color:#2f81f7; font-size:22px; font-weight:bold; margin-bottom:4px;")
        lay.addWidget(hdr)

        sub = QLabel("Automatizacion de Agendas y Gestion de Historias Clinicas")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub.setStyleSheet("color:#8b949e; font-size:13px; margin-bottom:12px;")
        lay.addWidget(sub)

        cards = QHBoxLayout()
        for titulo, val, col in [
            ("Version", VERSION, "#2f81f7"),
            ("Estado",  "Listo", "#3fb950"),
        ]:
            fr = QFrame()
            fr.setStyleSheet(f"background:#161b22; border:1px solid {col}; border-radius:8px; padding:4px;")
            cv = QVBoxLayout(fr)
            lv = QLabel(val)
            lv.setStyleSheet(f"color:{col}; font-size:20px; font-weight:bold;")
            lv.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lt = QLabel(titulo)
            lt.setStyleSheet("color:#8b949e; font-size:11px;")
            lt.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cv.addWidget(lv); cv.addWidget(lt)
            cards.addWidget(fr)
        lay.addLayout(cards)

        grp = QGroupBox("Actividad Reciente")
        lay.addWidget(grp)
        gl = QVBoxLayout(grp)
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setMaximumHeight(130)
        self.log_box.setStyleSheet(
            "background:#0d1117; color:#3fb950; font-family:Consolas,monospace; font-size:11px;"
        )
        gl.addWidget(self.log_box)

        self._log("Sistema iniciado")
        lay.addStretch()

    def _log(self, msg):
        ts = QDateTime.currentDateTime().toString("hh:mm:ss")
        self.log_box.append(f"[{ts}] {msg}")
        self.log_box.moveCursor(QTextCursor.MoveOperation.End)


# ===========================================================================
# Tab: Crear Agenda
# ===========================================================================
class AgendaTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        grp = QGroupBox("Parametros de Agenda")
        lay.addWidget(grp)
        form = QFormLayout(grp)
        self.cmb_sede = QComboBox()
        self.cmb_sede.addItem("-- Seleccionar sede --")
        cfg = load_main_config()
        self.cmb_sede.addItems(sorted(cfg.get("sedes", [])))
        form.addRow("Sede:", self.cmb_sede)
        self.txt_apellido = QLineEdit()
        self.txt_apellido.setPlaceholderText("Primer apellido del profesional")
        form.addRow("Apellido:", self.txt_apellido)
        self.txt_nombre = QLineEdit()
        self.txt_nombre.setPlaceholderText("Primer nombre del profesional")
        form.addRow("Nombre:", self.txt_nombre)
        self.txt_fi = QLineEdit(); self.txt_fi.setPlaceholderText("DD/MM/YYYY")
        self.txt_ff = QLineEdit(); self.txt_ff.setPlaceholderText("DD/MM/YYYY")
        form.addRow("Fecha inicio:", self.txt_fi)
        form.addRow("Fecha fin:", self.txt_ff)
        rb = QHBoxLayout()
        btn = QPushButton("Iniciar Creacion de Agenda")
        btn.setProperty("class", "success")
        btn.clicked.connect(self._run)
        rb.addWidget(btn)
        lay.addLayout(rb)
        lay.addWidget(QLabel("Log:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setStyleSheet("background:#0d1117; color:#3fb950; font-family:Consolas,monospace; font-size:11px;")
        lay.addWidget(self.log)

    def _run(self):
        if not self.txt_apellido.text().strip():
            QMessageBox.warning(self, "Error", "Ingresa el apellido del profesional.")
            return
        self.log.append("[INFO] Iniciando creacion de agenda...")
        self.log.append(f"      Profesional: {self.txt_apellido.text()} {self.txt_nombre.text()}")
        self.log.append("[NOTA] Conectar con bot_engine.Bot para ejecucion real.")


# ===========================================================================
# Worker en hilo aparte para descarga masiva HC (no bloquea la UI)
# ===========================================================================
class _HCWorker(QThread):
    progress      = pyqtSignal(int, int, str, str, str)  # completed, total, detail, current_item, status
    log_message   = pyqtSignal(str)
    finished_ok   = pyqtSignal(bool, str)                # ok, info

    def __init__(self, req, stop_event, run_hc_job_func, parent=None):
        super().__init__(parent)
        self._req       = req
        self._stop      = stop_event
        self._run_hc_job = run_hc_job_func

    def stop(self):
        try:
            if self._stop is not None:
                self._stop.set()
        except Exception:
            pass

    def run(self):
        def _cb(*args, **kwargs):
            # El runner puede llamar de dos formas:
            #   - progress_callback({"completed":..,"total":..,...})  (dict posicional)
            #   - progress_callback(completed=..., total=..., ...)    (kwargs)
            data = {}
            if args and isinstance(args[0], dict):
                data = dict(args[0])
            data.update(kwargs)
            try:
                completed    = int(data.get("completed") or 0)
                total        = int(data.get("total") or 0)
                detail       = str(data.get("detail") or "")
                current_item = str(data.get("current_item") or "")
                status       = str(data.get("status") or "running")
                self.progress.emit(completed, total, detail, current_item, status)
            except Exception as ex:
                # Loguear para que no se pierda silenciosamente
                try:
                    self.log_message.emit(f"[CB ERR] {ex} :: data={data}")
                except Exception:
                    pass

        try:
            self.log_message.emit("[BOT] Iniciando motor de descarga...")
            resultado = self._run_hc_job(self._req, stop_event=self._stop, progress_callback=_cb)
            info = ""
            try:
                if isinstance(resultado, dict):
                    info = (f"Procesados: {resultado.get('procesados', '?')} | "
                            f"Exitosos: {resultado.get('exitosos', '?')} | "
                            f"Fallidos: {resultado.get('fallidos', '?')}")
                else:
                    info = str(resultado or "")
            except Exception:
                info = ""
            self.finished_ok.emit(True, info)
        except Exception as ex:
            import traceback as _tb
            detalle = "".join(_tb.format_exception(type(ex), ex, ex.__traceback__))
            self.log_message.emit(f"[ERROR] {ex}")
            ruta = _escribir_error_txt("Error durante la descarga de HC", detalle)
            if ruta:
                self.log_message.emit(f"[ERROR] Detalle exacto guardado en: {ruta}")
            self.finished_ok.emit(False, str(ex))


# ===========================================================================
# Tab: Historia Clinica
# ===========================================================================
class HCTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._ids: List[str] = []
        self._rows: List[dict] = []   # filas completas con fechas por paciente
        self._worker: Optional[_HCWorker] = None
        self._stop_event = None
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        # ── Fechas compartidas ───────────────────────────────────────────
        grp_fecha = QGroupBox("Rango de Fechas (global — opcional si el Excel ya trae fechas por paciente)")
        lay.addWidget(grp_fecha)
        flay = QHBoxLayout(grp_fecha)
        flay.addWidget(QLabel("Fecha inicio:"))
        self.txt_fi = QLineEdit(); self.txt_fi.setPlaceholderText("DD/MM/YYYY"); self.txt_fi.setMaximumWidth(130)
        flay.addWidget(self.txt_fi)
        flay.addWidget(QLabel("Fecha fin:"))
        self.txt_ff = QLineEdit(); self.txt_ff.setPlaceholderText("DD/MM/YYYY"); self.txt_ff.setMaximumWidth(130)
        flay.addWidget(self.txt_ff)
        flay.addStretch()

        # ── Tabs: Individual / Masiva ────────────────────────────────────
        inner_tabs = QTabWidget()
        lay.addWidget(inner_tabs)

        # — Tab Individual —
        w_ind = QWidget()
        inner_tabs.addTab(w_ind, "Individual")
        vind = QVBoxLayout(w_ind)
        form = QFormLayout()
        self.txt_id = QLineEdit(); self.txt_id.setPlaceholderText("ID del paciente")
        form.addRow("ID Paciente:", self.txt_id)
        vind.addLayout(form)
        rb_ind = QHBoxLayout()
        btn_ind = QPushButton("Descargar HC")
        btn_ind.setProperty("class", "success")
        btn_ind.clicked.connect(self._run_individual)
        rb_ind.addWidget(btn_ind)
        btn_open = QPushButton("Abrir Carpeta")
        btn_open.clicked.connect(self._open)
        rb_ind.addWidget(btn_open)
        vind.addLayout(rb_ind)
        vind.addStretch()

        # — Tab Masiva —
        w_mas = QWidget()
        inner_tabs.addTab(w_mas, "Descarga Masiva")
        vmas = QVBoxLayout(w_mas)

        # Cargar archivo
        hf = QHBoxLayout()
        self.txt_archivo = QLineEdit(); self.txt_archivo.setPlaceholderText("Selecciona archivo Excel (.xlsx) o CSV con columna 'id' o 'documento'")
        self.txt_archivo.setReadOnly(True)
        btn_browse = QPushButton("Examinar...")
        btn_browse.setMaximumWidth(100)
        btn_browse.clicked.connect(self._browse_file)
        hf.addWidget(self.txt_archivo); hf.addWidget(btn_browse)
        vmas.addLayout(hf)

        # Tabla de IDs
        self.tabla = QTableWidget(0, 7)
        self.tabla.setHorizontalHeaderLabels(["#", "ID / Documento", "Servicio", "Fecha Inicio", "Fecha Fin", "Estado", "Acción"])
        self.tabla.horizontalHeader().setStretchLastSection(False)
        self.tabla.setColumnWidth(0, 40)
        self.tabla.setColumnWidth(1, 130)
        self.tabla.setColumnWidth(2, 200)
        self.tabla.setColumnWidth(3, 95)
        self.tabla.setColumnWidth(4, 95)
        self.tabla.setColumnWidth(5, 180)
        self.tabla.setColumnWidth(6, 80)
        # Estado por fila para contadores fiables (no dependen del texto):
        # valores posibles: 'pendiente', 'con_soporte', 'sin_hc', 'fallido'
        self._row_status: list = []
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setAlternatingRowColors(True)
        # Scroll vertical SIEMPRE visible y desactivar auto-scroll para que el
        # usuario pueda revisar la lista mientras se procesa.
        self.tabla.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.tabla.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.tabla.setAutoScroll(False)
        self.tabla.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.tabla.setMinimumHeight(280)
        vmas.addWidget(self.tabla)

        # Contador
        self.lbl_conteo = QLabel("0 pacientes cargados")
        self.lbl_conteo.setStyleSheet("color:#8b949e; font-size:11px;")
        vmas.addWidget(self.lbl_conteo)

        # ── Dashboard de progreso (en vivo) ─────────────────────────────
        grp_dash = QGroupBox("📊 Estado de la Descarga")
        grp_dash.setStyleSheet("QGroupBox { color:#e6edf3; font-weight:bold; }")
        vmas.addWidget(grp_dash)
        from PyQt6.QtWidgets import QGridLayout
        dash_lay = QGridLayout(grp_dash)
        dash_lay.setHorizontalSpacing(8)
        dash_lay.setVerticalSpacing(8)
        self._stats_labels: dict = {}
        cards_def = [
            ("total",       "Total",       "#2f81f7"),
            ("completados", "Con soporte", "#22c55e"),
            ("sin_hc",      "Sin HC",      "#eab308"),
            ("fallidos",    "Fallidos",    "#ef4444"),
            ("validados",   "PDFs OK",     "#2f81f7"),
            ("restantes",   "Pendientes",  "#f59e0b"),
            ("actual",      "Actual",      "#06b6d4"),
        ]
        # Distribuir en grilla 4 col (responsive: se acomoda al re-dimensionar)
        cols = 4
        for idx, (key, titulo, color) in enumerate(cards_def):
            fr = QFrame()
            fr.setStyleSheet(
                f"background:#0d1117; border:1px solid {color}; border-radius:8px; padding:6px;"
            )
            fr.setMinimumWidth(110)
            fr.setMinimumHeight(70)
            fr.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            v = QVBoxLayout(fr)
            v.setContentsMargins(4, 4, 4, 4)
            v.setSpacing(2)
            lv = QLabel("0" if key != "actual" else "—")
            lv.setStyleSheet(f"color:{color}; font-size:16px; font-weight:bold;")
            lv.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lv.setWordWrap(True)
            lt = QLabel(titulo)
            lt.setStyleSheet("color:#8b949e; font-size:10px;")
            lt.setAlignment(Qt.AlignmentFlag.AlignCenter)
            v.addWidget(lv); v.addWidget(lt)
            self._stats_labels[key] = lv
            r, c = divmod(idx, cols)
            dash_lay.addWidget(fr, r, c)
        # Hacer que las columnas se expandan equitativamente
        for c in range(cols):
            dash_lay.setColumnStretch(c, 1)

        # Progress bar
        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.progress.setTextVisible(True)
        self.progress.setFormat("%v / %m pacientes")
        vmas.addWidget(self.progress)

        # Botones masiva
        rb_mas = QHBoxLayout()
        btn_masiva = QPushButton("Iniciar Descarga Masiva")
        btn_masiva.setProperty("class", "success")
        btn_masiva.clicked.connect(self._run_masiva)
        self.btn_detener = QPushButton("Detener")
        self.btn_detener.setProperty("class", "danger")
        self.btn_detener.clicked.connect(self._detener_masiva)
        self.btn_continuar = QPushButton("▶️ Continuar")
        self.btn_continuar.setProperty("class", "success")
        self.btn_continuar.setToolTip(
            "Reanuda la descarga desde el primer paciente PENDIENTE o FALLIDO,\n"
            "omitiendo los que ya estan 'Con soporte'."
        )
        self.btn_continuar.clicked.connect(self._continuar_masiva)
        self.btn_continuar.setEnabled(False)
        btn_validar = QPushButton("🔍 Validar PDFs descargados")
        btn_validar.setToolTip(
            "Escanea la carpeta de descargas y verifica que cada PDF tenga "
            "contenido y que la fecha/servicio coincidan con lo esperado."
        )
        btn_validar.clicked.connect(self._validar_descargas)
        btn_limpiar = QPushButton("Limpiar Lista")
        btn_limpiar.clicked.connect(self._limpiar)
        btn_open2 = QPushButton("Abrir Carpeta")
        btn_open2.clicked.connect(self._open)
        # Botón "Mantener equipo despierto" (anti-bloqueo Windows)
        self.btn_keep_awake = QPushButton("☕ Mantener PC despierta")
        self.btn_keep_awake.setCheckable(True)
        self.btn_keep_awake.setToolTip(
            "Evita que Windows entre en suspensión, apague la pantalla o se "
            "bloquee mientras el bot está corriendo. Click para activar / desactivar."
        )
        self.btn_keep_awake.toggled.connect(self._toggle_keep_awake)
        rb_mas.addWidget(btn_masiva); rb_mas.addWidget(self.btn_detener)
        rb_mas.addWidget(self.btn_continuar)
        rb_mas.addWidget(btn_validar)
        rb_mas.addWidget(btn_limpiar); rb_mas.addWidget(btn_open2)
        rb_mas.addWidget(self.btn_keep_awake)
        vmas.addLayout(rb_mas)

        # Log
        lay.addWidget(QLabel("Log:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setStyleSheet("background:#0d1117; color:#3fb950; font-family:Consolas,monospace; font-size:11px;")
        self.log.setMaximumHeight(130)
        lay.addWidget(self.log)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo de pacientes",
            str(Path.home()),
            "Archivos de datos (*.xlsx *.xls *.csv);;Todos (*.*)"
        )
        if not path:
            return
        self.txt_archivo.setText(path)
        self._cargar_ids(path)

    def _cargar_ids(self, path: str):
        self._ids = []
        self._rows = []
        self.tabla.setRowCount(0)
        try:
            p = Path(path)
            if p.suffix.lower() in (".xlsx", ".xls"):
                try:
                    import pandas as pd
                    df = pd.read_excel(path, dtype=str, engine="openpyxl")
                except Exception as ex_xlsx:
                    detalle = str(ex_xlsx)
                    self.log.append(f"[ERROR] No se pudo leer el Excel: {detalle}")
                    _escribir_error_txt("Error al leer Excel de pacientes", detalle)
                    QMessageBox.critical(
                        self, "Error al leer Excel",
                        f"No se pudo leer el archivo Excel.\n\nDetalle:\n{detalle}\n\n"
                        "Sugerencia: vuelve a guardar el archivo como .xlsx, o "
                        "exportalo a CSV y cargalo asi.")
                    return
            else:
                try:
                    import pandas as pd
                    df = pd.read_csv(path, dtype=str, encoding="utf-8")
                except Exception:
                    try:
                        import pandas as pd
                        df = pd.read_csv(path, dtype=str, encoding="latin-1")
                    except Exception:
                        self.log.append("[ERROR] No se pudo leer el CSV.")
                        return

            # Normalizar nombres de columnas para búsqueda insensible
            col_map = {c.strip().upper(): c for c in df.columns}

            # Buscar columna con IDs
            col = None
            for c in df.columns:
                if c.strip().lower() in ("cc", "id", "documento", "cedula", "nro_documento", "id_paciente", "paciente"):
                    col = c
                    break
            if col is None:
                col = df.columns[0]
                self.log.append(f"[AVISO] Columna 'id' no encontrada, usando primera columna: '{col}'")

            # Buscar columnas opcionales del Excel
            def _find_col(*names):
                for n in names:
                    if n.upper() in col_map:
                        return col_map[n.upper()]
                return None

            col_fi   = _find_col("FECHA INICIO", "FECHA_INICIO", "F_INICIO", "DESDE")
            col_ff   = _find_col("FECHA FIN", "FECHA_FIN", "F_FIN", "HASTA")
            col_srv  = _find_col("SERVICIO")
            col_est  = _find_col("ESTRATEGIA")
            col_fac  = _find_col("NUMERO DE FACTURA", "NUMERO_FACTURA", "FACTURA", "NRO_FACTURA")
            col_ing  = _find_col("NUMERO DE INGRESO", "NUMERO INGRESO", "NUMERO_INGRESO", "NRO INGRESO", "NRO_INGRESO", "INGRESO")

            tiene_fechas = bool(col_fi and col_ff)
            if tiene_fechas:
                self.log.append(f"[OK] Columnas FECHA INICIO / FECHA FIN detectadas — se usarán fechas del Excel")
            else:
                self.log.append(f"[INFO] El Excel no tiene columnas de fecha — ingresa el rango global arriba")

            # Construir lista de filas. Deduplica por la combinacion
            # (cedula + servicio + fecha_inicio + fecha_fin) para que una misma
            # CC con varios servicios o varios rangos NO se pierda.
            seen = set()
            rows = []
            for _, row in df.iterrows():
                id_val = str(row[col]).strip() if row[col] == row[col] else ""
                if not id_val or id_val.lower() in ("nan", "none", ""):
                    continue
                fi_val  = str(row[col_fi]).strip()  if col_fi  and row[col_fi]  == row[col_fi]  else ""
                ff_val  = str(row[col_ff]).strip()  if col_ff  and row[col_ff]  == row[col_ff]  else ""
                srv_val = str(row[col_srv]).strip() if col_srv and row[col_srv] == row[col_srv] else ""
                est_val = str(row[col_est]).strip() if col_est and row[col_est] == row[col_est] else ""
                fac_val = str(row[col_fac]).strip() if col_fac and row[col_fac] == row[col_fac] else ""
                ing_val = str(row[col_ing]).strip() if col_ing and row[col_ing] == row[col_ing] else ""
                if ing_val.endswith(".0"):
                    ing_val = ing_val[:-2]
                if ing_val.lower() in ("nan", "none"):
                    ing_val = ""

                clave = (id_val, srv_val.upper(), fi_val, ff_val, fac_val.upper())
                if clave in seen:
                    continue
                seen.add(clave)
                rows.append({
                    "id":           id_val,
                    "fecha_inicio": fi_val,
                    "fecha_fin":    ff_val,
                    "servicio":     srv_val,
                    "estrategia":   est_val,
                    "factura":      fac_val,
                    "ingreso":      ing_val,
                })

            self._rows = rows
            self._ids  = [r["id"] for r in rows]
            self._row_status = ["pendiente"] * len(rows)

            self.tabla.setRowCount(len(rows))
            for i, r in enumerate(rows):
                self.tabla.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                self.tabla.setItem(i, 1, QTableWidgetItem(r["id"]))
                self.tabla.setItem(i, 2, QTableWidgetItem(r["servicio"]      or "—"))
                self.tabla.setItem(i, 3, QTableWidgetItem(r["fecha_inicio"] or "—"))
                self.tabla.setItem(i, 4, QTableWidgetItem(r["fecha_fin"]    or "—"))
                self.tabla.setItem(i, 5, QTableWidgetItem("Pendiente"))
                # Si ya hay un PDF descargado para esa CC en la carpeta destino,
                # mostrar el boton Ver desde el inicio.
                if self._buscar_pdf_de_fila(r["id"], r.get("servicio") or ""):
                    self._set_boton_ver(i, r["id"], r.get("servicio") or "")

            self.progress.setMaximum(len(rows))
            self.progress.setValue(0)
            self.lbl_conteo.setText(f"{len(rows)} pacientes cargados desde '{Path(path).name}'")
            self.log.append(f"[OK] {len(rows)} IDs cargados desde {Path(path).name}")

        except Exception as e:
            self.log.append(f"[ERROR] {e}")

    def _limpiar(self):
        self._ids = []
        self._rows = []
        self.tabla.setRowCount(0)
        self.txt_archivo.clear()
        self.progress.setValue(0)
        self.lbl_conteo.setText("0 pacientes cargados")
        self.log.append("[INFO] Lista limpiada.")

    def _run_individual(self):
        if not self.txt_id.text().strip():
            QMessageBox.warning(self, "Error", "Ingresa el ID del paciente.")
            return
        fi = self.txt_fi.text().strip()
        ff = self.txt_ff.text().strip()
        self.log.append(f"[INFO] Descargando HC paciente {self.txt_id.text()} | {fi} → {ff}")
        self.log.append("[NOTA] Conectar con bot_engine.Bot para ejecucion real.")

    def _run_masiva(self):
        if not self._rows:
            QMessageBox.warning(self, "Sin datos", "Primero carga un archivo con IDs de pacientes.")
            return

        # Worker ya activo
        if getattr(self, "_worker", None) and self._worker.isRunning():
            QMessageBox.information(self, "En ejecución",
                "Ya hay una descarga masiva en curso. Espera a que termine o usa Detener.")
            return

        fi_global = self.txt_fi.text().strip()
        ff_global = self.txt_ff.text().strip()

        tiene_fechas_excel = any(r["fecha_inicio"] and r["fecha_fin"] for r in self._rows)
        if not tiene_fechas_excel and (not fi_global or not ff_global):
            QMessageBox.warning(self, "Fechas requeridas",
                "El archivo no contiene columnas FECHA INICIO / FECHA FIN.\n"
                "Ingresa el rango de fechas global arriba.")
            return

        # Cargar runner del orquestador
        try:
            from agenda_app.bot.runner import run_hc_job
            from agenda_app.domain.requests import HCRequest
        except Exception as ex:
            QMessageBox.critical(self, "Error de importación",
                f"No se pudo cargar el motor del bot:\n{ex}")
            return

        # Construir lista de pacientes en formato esperado por el runner
        # IMPORTANTE: las claves DEBEN coincidir con las que lee runner.py
        # (fecha_inicio, fecha_fin, numero_factura) o el bot ignorara las
        # fechas por paciente y usara las del UI global.
        pacientes = []
        for r in self._rows:
            f_ini = r["fecha_inicio"] or fi_global
            f_fin = r["fecha_fin"] or ff_global
            factura = r["factura"] or ""
            ingreso = r.get("ingreso") or ""
            pacientes.append({
                "cedula":              r["id"],
                "servicio":            r["servicio"]    or "",
                "estrategia":          r["estrategia"]  or "RECIENTE",
                "fecha_inicio":        f_ini,
                "fecha_fin":           f_fin,
                "numero_factura":      factura,
                "numero_ingreso":      ingreso,
                # Compatibilidad con codigos legacy que aun lean estas claves:
                "FECHA INICIO":        f_ini,
                "FECHA FIN":           f_fin,
                "NUMERO DE FACTURA":   factura,
                "NUMERO DE INGRESO":   ingreso,
            })

        try:
            req = HCRequest(
                cedula="",
                fecha_inicio=fi_global or "",
                fecha_fin=ff_global or "",
                servicio="Todos",
                estrategia="RECIENTE",
                ruta_descarga="",
                pacientes_excel=pacientes,
            )
        except Exception as ex:
            QMessageBox.critical(self, "Error", f"No se pudo construir la petición:\n{ex}")
            return

        # Reset estado UI + dashboard
        self._row_status = ["pendiente"] * len(self._rows)
        for i in range(len(self._rows)):
            self.tabla.setItem(i, 5, QTableWidgetItem("Pendiente"))
            # Limpiar boton Ver previo
            self.tabla.removeCellWidget(i, 6)
        total = len(self._rows)
        self.progress.setValue(0)
        self.progress.setMaximum(total)
        self._stats_labels["total"].setText(str(total))
        self._stats_labels["completados"].setText("0")
        self._stats_labels["fallidos"].setText("0")
        self._stats_labels["validados"].setText("0")
        self._stats_labels["restantes"].setText(str(total))
        self._stats_labels["actual"].setText("—")
        self.log.append(f"[INICIO] Descarga masiva: {total} pacientes")

        # Lanzar worker
        self._stop_event = __import__("threading").Event()
        self._worker = _HCWorker(req, self._stop_event, run_hc_job)
        self._worker.progress.connect(self._on_progress)
        self._worker.log_message.connect(lambda msg: self.log.append(msg))
        self._worker.finished_ok.connect(self._on_finished)
        # Auto-activar 'Mantener PC despierta' al iniciar la descarga masiva
        try:
            if hasattr(self, "btn_keep_awake") and not self.btn_keep_awake.isChecked():
                self.btn_keep_awake.setChecked(True)
        except Exception:
            pass
        self._worker.start()

    def _parse_current_item(self, current_item: str):
        """current_item viene como 'CC <cedula> - <SERVICIO>'. Retorna (cedula, servicio)."""
        s = (current_item or "").strip()
        if not s:
            return "", ""
        if s.upper().startswith("CC "):
            s = s[3:].strip()
        if " - " in s:
            ced, srv = s.split(" - ", 1)
            return ced.strip(), srv.strip()
        return s, ""

    def _carpeta_descargas_actual(self) -> str:
        """Devuelve la carpeta donde el runner guarda los lotes HC_MASIVA_*."""
        try:
            with open(_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            base = str((cfg.get("navegadores") or {}).get("carpeta_descargas", "")).strip()
        except Exception:
            base = ""
        if base and Path(base).exists():
            return base
        # Fallback: carpeta downloads junto al exe
        fallback = _DATA_DIR / "downloads"
        try:
            fallback.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        return str(fallback)

    def _buscar_pdf_de_fila(self, cedula: str, servicio: str):
        """Busca recursivamente un PDF cuyo nombre empiece por <CC>_."""
        try:
            base = self._carpeta_descargas_actual()
            if not base or not Path(base).exists():
                return None
            cc = str(cedula).strip()
            if not cc:
                return None
            # Priorizar el lote mas reciente (HC_MASIVA_*)
            lotes = sorted(
                [d for d in Path(base).iterdir() if d.is_dir() and d.name.startswith("HC_MASIVA_")],
                key=lambda d: d.stat().st_mtime, reverse=True,
            )
            busquedas = lotes + [Path(base)]
            for raiz in busquedas:
                for pdf in raiz.rglob(f"{cc}_*.pdf"):
                    return str(pdf)
            return None
        except Exception:
            return None

    def _abrir_pdf_fila(self, cedula: str, servicio: str):
        ruta = self._buscar_pdf_de_fila(cedula, servicio)
        if not ruta:
            QMessageBox.information(self, "PDF no encontrado",
                f"No se encontro un PDF descargado para CC={cedula}.\n"
                f"Es posible que aun no se haya guardado o que la carpeta de "
                f"descargas haya cambiado.")
            return
        try:
            os.startfile(ruta)  # type: ignore[attr-defined]
        except Exception as ex:
            QMessageBox.warning(self, "No se pudo abrir", f"{ruta}\n\n{ex}")

    def _set_boton_ver(self, fila_idx: int, cedula: str, servicio: str):
        btn = QPushButton("📄 Ver")
        btn.setToolTip(f"Abrir PDF descargado de CC {cedula}")
        btn.clicked.connect(lambda _=False, c=cedula, s=servicio: self._abrir_pdf_fila(c, s))
        self.tabla.setCellWidget(fila_idx, 6, btn)

    def _on_progress(self, completed, total, detail, current_item, status):
        # Parsear current_item "CC <cedula> - <SERVICIO>"
        cedula_evt, servicio_evt = self._parse_current_item(current_item)
        estado_lower = (status or "").strip().lower()
        detalle_lower = (detail or "").lower()

        # Clasificar el resultado del evento en 3 estados claros:
        #   - 'con_soporte' : descarga exitosa con PDF guardado
        #   - 'sin_hc'      : el paciente no tiene HC en el rango (resultado de negocio)
        #   - 'fallido'     : error tecnico
        es_sin_hc = any(k in detalle_lower for k in (
            "sin hc", "sin registros", "no encontr", "no hay resultados",
            "sin resultados", "no tiene hc",
        ))
        es_con_soporte = (
            estado_lower in ("ok", "success", "exitoso", "descargado", "completado", "completed")
            or any(k in detalle_lower for k in (
                "descargado correctamente", "tiene hc y ya estaba", "exito",
                "archivo organizado", "guardado",
            ))
        ) and not es_sin_hc
        es_fallo = (
            not es_con_soporte and not es_sin_hc and (
                estado_lower in ("error", "fallido", "failed", "fail")
                or any(k in detalle_lower for k in ("error", "fall", "rechazado"))
            )
        )

        # Buscar fila correspondiente: id == cedula y (servicio coincide o vacio)
        fila_match = -1
        if cedula_evt:
            srv_norm = servicio_evt.upper().strip()
            for i, r in enumerate(self._rows):
                if r["id"] != cedula_evt:
                    continue
                r_srv = (r.get("servicio") or "").upper().strip()
                if not srv_norm or not r_srv or srv_norm == r_srv or srv_norm in r_srv or r_srv in srv_norm:
                    fila_match = i
                    break

        if fila_match >= 0:
            if es_con_soporte:
                item = QTableWidgetItem("✅ CON SOPORTE")
                item.setForeground(Qt.GlobalColor.green)
                if fila_match < len(self._row_status):
                    self._row_status[fila_match] = "con_soporte"
                self._set_boton_ver(fila_match, cedula_evt, servicio_evt)
            elif es_sin_hc:
                item = QTableWidgetItem("⚠ SIN HC")
                item.setForeground(Qt.GlobalColor.yellow)
                if fila_match < len(self._row_status):
                    self._row_status[fila_match] = "sin_hc"
            elif es_fallo:
                item = QTableWidgetItem(f"❌ {(detail or status or 'ERROR').strip()[:60]}")
                item.setForeground(Qt.GlobalColor.red)
                if fila_match < len(self._row_status):
                    self._row_status[fila_match] = "fallido"
            else:
                # Evento intermedio "Procesando..." - solo actualizar texto sin cambiar status
                txt = (detail or status or "Procesando...").strip()[:80]
                item = QTableWidgetItem(txt)
                item.setForeground(Qt.GlobalColor.cyan)
            self.tabla.setItem(fila_match, 5, item)
            # NO hacer scrollToItem: respetar el scroll manual del usuario.

        if cedula_evt:
            self._stats_labels["actual"].setText(cedula_evt)

        # Barra de progreso
        if total and total > 0:
            self.progress.setMaximum(total)
            self.progress.setValue(completed)

        # Contadores con base en _row_status (fiable)
        ok = sum(1 for s in self._row_status if s == "con_soporte")
        sinhc = sum(1 for s in self._row_status if s == "sin_hc")
        falli = sum(1 for s in self._row_status if s == "fallido")
        pend = sum(1 for s in self._row_status if s == "pendiente")
        self._stats_labels["completados"].setText(str(ok))
        self._stats_labels["fallidos"].setText(str(falli))
        if "sin_hc" in self._stats_labels:
            self._stats_labels["sin_hc"].setText(str(sinhc))
        self._stats_labels["restantes"].setText(str(pend))

    def _on_finished(self, ok, info):
        self._stats_labels["actual"].setText("—")
        # Resetear estado del boton Detener
        try:
            self.btn_detener.setText("Detener")
            self.btn_detener.setEnabled(True)
            self._stop_click_count = 0
        except Exception:
            pass
        # Liberar el bloqueo de suspensión al terminar
        try:
            if hasattr(self, "btn_keep_awake") and self.btn_keep_awake.isChecked():
                self.btn_keep_awake.setChecked(False)
        except Exception:
            pass
        if ok:
            self.log.append(f"[FIN] Descarga masiva completada: {info}")
            QMessageBox.information(self, "Completado",
                f"Descarga masiva finalizada.\n{info}\n\n"
                "Pulsa '🔍 Validar PDFs descargados' para verificar "
                "contenido / fechas / servicio de cada PDF.")
        else:
            self.log.append(f"[FIN] Descarga masiva con errores: {info}")
            QMessageBox.warning(self, "Finalizado con errores",
                f"La descarga terminó con problemas:\n{info}")

    def _detener_masiva(self):
        if not self._worker or not self._worker.isRunning():
            # Reset visual por si quedo en estado intermedio
            try:
                self.btn_detener.setText("Detener")
                self.btn_detener.setEnabled(True)
                self._stop_click_count = 0
            except Exception:
                pass
            QMessageBox.information(self, "Sin proceso", "No hay descarga masiva en curso.")
            return

        # Escalado: 1er click = soft stop, 2do click = forzar terminacion del hilo.
        if not hasattr(self, "_stop_click_count"):
            self._stop_click_count = 0
        self._stop_click_count += 1

        if self._stop_click_count == 1:
            self.log.append("[STOP] Solicitando detener (rompe sub-loops y termina tras tarea actual). Pulsa de nuevo para FORZAR cierre.")
            try:
                self._worker.stop()
            except Exception as ex:
                self.log.append(f"[STOP] Error señalando detener: {ex}")
            try:
                self.btn_detener.setText("Forzar parada")
                self.btn_continuar.setEnabled(True)
            except Exception:
                pass
            return

        # Segundo click: forzar
        confirmar = QMessageBox.question(
            self, "Forzar parada",
            "Esto matará el hilo de descarga inmediatamente.\n"
            "Puede dejar el navegador abierto y archivos parciales.\n\n¿Continuar?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirmar != QMessageBox.StandardButton.Yes:
            return
        self.log.append("[STOP] FORZANDO terminación del hilo de descarga...")
        try:
            self._worker.requestInterruption()
        except Exception:
            pass
        try:
            self._worker.terminate()
            self._worker.wait(3000)
        except Exception as ex:
            self.log.append(f"[STOP] Error forzando: {ex}")
        try:
            self.btn_detener.setText("Detener")
            self.btn_detener.setEnabled(True)
            self.btn_continuar.setEnabled(True)
            self._stop_click_count = 0
        except Exception:
            pass
        self.log.append("[STOP] Hilo terminado a la fuerza. Puedes pulsar Continuar para reanudar.")

    def _continuar_masiva(self):
        """Reanuda la descarga masiva tomando solo los pacientes que aun NO
        tienen estado 'con_soporte'. Util tras pulsar Detener o tras un cierre.
        """
        if self._worker and self._worker.isRunning():
            QMessageBox.information(self, "En ejecución",
                "Ya hay una descarga en curso. Pulsa Detener primero.")
            return
        if not self._rows:
            QMessageBox.warning(self, "Sin datos", "Primero carga un archivo.")
            return
        if not self._row_status or len(self._row_status) != len(self._rows):
            self._row_status = ["pendiente"] * len(self._rows)

        # Filtrar pacientes pendientes / fallidos / sin_hc (no descargados ok).
        indices_pendientes = [i for i, s in enumerate(self._row_status)
                              if s in ("pendiente", "fallido", "sin_hc")]
        if not indices_pendientes:
            QMessageBox.information(self, "Nada que continuar",
                "Todos los pacientes ya estan 'Con soporte'.")
            return

        fi_global = self.txt_fi.text().strip()
        ff_global = self.txt_ff.text().strip()
        try:
            from agenda_app.bot.runner import run_hc_job
            from agenda_app.domain.requests import HCRequest
        except Exception as ex:
            QMessageBox.critical(self, "Error de importación",
                f"No se pudo cargar el motor del bot:\n{ex}")
            return

        pacientes = []
        for i in indices_pendientes:
            r = self._rows[i]
            f_ini = r["fecha_inicio"] or fi_global
            f_fin = r["fecha_fin"] or ff_global
            factura = r["factura"] or ""
            ingreso = r.get("ingreso") or ""
            pacientes.append({
                "cedula":              r["id"],
                "servicio":            r["servicio"]    or "",
                "estrategia":          r["estrategia"]  or "RECIENTE",
                "fecha_inicio":        f_ini,
                "fecha_fin":           f_fin,
                "numero_factura":      factura,
                "numero_ingreso":      ingreso,
                "FECHA INICIO":        f_ini,
                "FECHA FIN":           f_fin,
                "NUMERO DE FACTURA":   factura,
                "NUMERO DE INGRESO":   ingreso,
            })

        try:
            req = HCRequest(
                cedula="",
                fecha_inicio=fi_global or "",
                fecha_fin=ff_global or "",
                servicio="Todos",
                estrategia="RECIENTE",
                ruta_descarga="",
                pacientes_excel=pacientes,
            )
        except Exception as ex:
            QMessageBox.critical(self, "Error", f"No se pudo construir la petición:\n{ex}")
            return

        # NO reseteamos _row_status (preservamos los que ya estan con_soporte).
        # Marcamos los reanudados como pendiente para que el progreso refleje bien.
        for i in indices_pendientes:
            self._row_status[i] = "pendiente"
            self.tabla.setItem(i, 5, QTableWidgetItem("Pendiente"))

        total_global = len(self._rows)
        completados_global = sum(1 for s in self._row_status if s == "con_soporte")
        self.progress.setMaximum(total_global)
        self.progress.setValue(completados_global)
        self.log.append(f"[CONTINUAR] Reanudando {len(indices_pendientes)} pacientes pendientes "
                        f"(ya completados: {completados_global}/{total_global})")

        self._stop_event = __import__("threading").Event()
        self._worker = _HCWorker(req, self._stop_event, run_hc_job)
        self._worker.progress.connect(self._on_progress)
        self._worker.log_message.connect(lambda msg: self.log.append(msg))
        self._worker.finished_ok.connect(self._on_finished)
        try:
            if hasattr(self, "btn_keep_awake") and not self.btn_keep_awake.isChecked():
                self.btn_keep_awake.setChecked(True)
        except Exception:
            pass
        self.btn_continuar.setEnabled(False)
        self._worker.start()

    def _open(self):
        d = _ROOT_DIR / "downloads" / "hc"
        d.mkdir(parents=True, exist_ok=True)
        os.startfile(str(d))

    # ------------------------------------------------------------------
    # Mantener PC despierta (anti-bloqueo de Windows durante descarga)
    # ------------------------------------------------------------------
    def _toggle_keep_awake(self, activo: bool):
        """Activa/desactiva el modo 'no dormir' usando Windows SetThreadExecutionState.
        Mientras está activo, Windows no entrará en suspensión, no apagará la
        pantalla y no bloqueará la sesión por inactividad.
        """
        try:
            import ctypes
            ES_CONTINUOUS       = 0x80000000
            ES_SYSTEM_REQUIRED  = 0x00000001
            ES_DISPLAY_REQUIRED = 0x00000002
            ES_AWAYMODE_REQUIRED= 0x00000040
            if activo:
                # Mantener sistema + pantalla activos
                flags = ES_CONTINUOUS | ES_SYSTEM_REQUIRED | ES_DISPLAY_REQUIRED | ES_AWAYMODE_REQUIRED
                rc = ctypes.windll.kernel32.SetThreadExecutionState(flags)
                if rc == 0:
                    self.log.append("[KEEP-AWAKE] No se pudo activar (API devolvio 0).")
                    self.btn_keep_awake.setChecked(False)
                    return
                self.btn_keep_awake.setText("☀ PC despierta (ACTIVO)")
                self.btn_keep_awake.setStyleSheet("background:#16a34a; color:white; font-weight:bold;")
                self.log.append("[KEEP-AWAKE] Activado. Windows no se suspendera durante la descarga.")
            else:
                ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS)
                self.btn_keep_awake.setText("☕ Mantener PC despierta")
                self.btn_keep_awake.setStyleSheet("")
                self.log.append("[KEEP-AWAKE] Desactivado.")
        except Exception as ex:
            self.log.append(f"[KEEP-AWAKE] Error: {ex}")
            try:
                self.btn_keep_awake.setChecked(False)
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Validación de PDFs ya descargados (sin conectar al bot)
    # ------------------------------------------------------------------
    def _validar_descargas(self):
        """Escanea downloads/hc y verifica cada PDF contra los datos cargados:
        - Que el PDF no esté vacío (mínimo de bytes y de texto)
        - Que la fecha extraída esté en el rango fi/ff del paciente
        - Que el servicio del PDF coincida con el esperado
        """
        if not self._rows:
            QMessageBox.warning(self, "Sin datos",
                "Primero carga el archivo de pacientes para saber qué validar.")
            return

        # Extraer texto con PyMuPDF
        try:
            import fitz  # PyMuPDF
        except Exception as ex:
            QMessageBox.critical(self, "Falta dependencia",
                f"No se pudo cargar PyMuPDF (fitz):\n{ex}")
            return

        import glob, re
        from datetime import datetime

        base_dl = _ROOT_DIR / "downloads" / "hc"
        if not base_dl.exists():
            QMessageBox.information(self, "Sin descargas",
                f"No existe la carpeta de descargas:\n{base_dl}")
            return

        PDF_MIN_BYTES = 5 * 1024
        PDF_MIN_TEXTO = 80

        def _parse_fecha(txt):
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return datetime.strptime(txt.strip(), fmt)
                except Exception:
                    pass
            return None

        def _normalizar(s):
            return re.sub(r"\s+", " ", (s or "").upper().strip())

        validados = 0
        invalidos = 0
        sin_archivo = 0
        self.log.append("")
        self.log.append("[VALIDAR] Escaneando PDFs descargados...")

        for i, r in enumerate(self._rows):
            cedula   = str(r["id"]).strip()
            servicio = _normalizar(r["servicio"])
            fi_txt   = r["fecha_inicio"] or self.txt_fi.text().strip()
            ff_txt   = r["fecha_fin"]    or self.txt_ff.text().strip()
            fi = _parse_fecha(fi_txt) if fi_txt else None
            ff = _parse_fecha(ff_txt) if ff_txt else None

            # Buscar PDFs del paciente en cualquier subcarpeta
            patrones = [
                str(base_dl / "**" / f"{cedula}*.pdf"),
                str(base_dl / "**" / f"*{cedula}*.pdf"),
            ]
            archivos = []
            for pat in patrones:
                archivos.extend(glob.glob(pat, recursive=True))
            archivos = list({a for a in archivos})

            if not archivos:
                sin_archivo += 1
                item = QTableWidgetItem("⚠ Sin PDF descargado")
                item.setForeground(Qt.GlobalColor.yellow)
                self.tabla.setItem(i, 5, item)
                continue

            # Validar el más reciente
            archivo = max(archivos, key=lambda a: os.path.getmtime(a))
            problemas = []
            try:
                tam = os.path.getsize(archivo)
                if tam < PDF_MIN_BYTES:
                    problemas.append(f"PDF muy pequeño ({tam}B)")
                with fitz.open(archivo) as doc:
                    texto = "\n".join(p.get_text() for p in doc)
                texto_norm = _normalizar(texto)
                if len(texto_norm) < PDF_MIN_TEXTO:
                    problemas.append("PDF sin texto / vacío")
                # Servicio
                if servicio and servicio not in texto_norm:
                    # tolerar primera palabra
                    primera = servicio.split()[0] if servicio else ""
                    if primera and primera not in texto_norm:
                        problemas.append(f"servicio '{r['servicio']}' no coincide")
                # Fecha en el texto
                if fi and ff:
                    fechas = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", texto)
                    parsed = [d for d in (_parse_fecha(f) for f in fechas) if d]
                    if parsed:
                        if not any(fi <= d <= ff for d in parsed):
                            problemas.append(
                                f"sin fecha en rango {fi.strftime('%d/%m/%Y')}—"
                                f"{ff.strftime('%d/%m/%Y')}"
                            )
            except Exception as ex:
                problemas.append(f"error abriendo PDF: {ex}")

            if problemas:
                invalidos += 1
                item = QTableWidgetItem("❌ " + "; ".join(problemas)[:80])
                item.setForeground(Qt.GlobalColor.red)
                item.setToolTip(archivo + "\n" + "\n".join(problemas))
                self.tabla.setItem(i, 5, item)
            else:
                validados += 1
                item = QTableWidgetItem("✅ PDF válido")
                item.setForeground(Qt.GlobalColor.green)
                item.setToolTip(archivo)
                self.tabla.setItem(i, 5, item)

        self._stats_labels["validados"].setText(str(validados))
        self.log.append(
            f"[VALIDAR] OK={validados}  FALLAS={invalidos}  SIN_PDF={sin_archivo}"
        )
        QMessageBox.information(self, "Validación completada",
            f"✅ PDFs válidos:    {validados}\n"
            f"❌ PDFs con fallas: {invalidos}\n"
            f"⚠  Sin PDF:         {sin_archivo}")


# ===========================================================================
# Ventana Principal
# ===========================================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"BOT  v{VERSION}  —  Panel de Control")
        # Minimo modesto para que entre en pantallas pequenas (1366x768) sin
        # recortar nada: el contenido SCROLLEA en vez de esconder botones.
        self.setMinimumSize(820, 560)
        self.resize(1120, 760)
        self.setStyleSheet(DARK_STYLE)
        self._build_ui()
        self._build_menu()
        sb = QStatusBar()
        self.setStatusBar(sb)
        sb.showMessage(f"BOT v{VERSION} — Listo")
        self.show()
        # Chequeo automatico de nueva version al iniciar (no molesta si esta al
        # dia; solo ofrece actualizar si hay una version mas reciente).
        try:
            QTimer.singleShot(3000, self._check_updates)
        except Exception:
            pass

    @staticmethod
    def _scrollable(widget):
        """Envuelve un widget en un area con scroll vertical para que, si la
        ventana es pequena, el contenido se desplace en vez de recortarse
        (evita que los botones se escondan o se distorsione el layout)."""
        sa = QScrollArea()
        sa.setWidgetResizable(True)
        sa.setFrameShape(QFrame.Shape.NoFrame)
        sa.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        sa.setWidget(widget)
        return sa

    def _build_ui(self):
        cw = QWidget()
        self.setCentralWidget(cw)
        vl = QVBoxLayout(cw)
        vl.setContentsMargins(10, 10, 10, 6)
        tabs = QTabWidget()
        tabs.setDocumentMode(True)
        vl.addWidget(tabs)
        self.tab_dash  = DashboardTab()
        self.tab_cred  = CredencialesTab()
        self.tab_nav   = NavegadoresTab()
        self.tab_hc    = HCTab()
        tabs.addTab(self._scrollable(self.tab_dash),  "Dashboard")
        tabs.addTab(self._scrollable(self.tab_cred),  "Credenciales")
        tabs.addTab(self._scrollable(self.tab_nav),   "Navegadores")
        tabs.addTab(self._scrollable(self.tab_hc),    "Historia Clinica")

    def _build_menu(self):
        mb = self.menuBar()
        # En Windows el dark style puede esconder/desactivar el menú nativo.
        # Forzamos QMenuBar widget-based para que siempre se vea y responda.
        try:
            mb.setNativeMenuBar(False)
        except Exception:
            pass
        mb.setStyleSheet("""
            QMenuBar { background:#161b22; color:#e6edf3; padding:4px; }
            QMenuBar::item { background:transparent; padding:6px 14px; }
            QMenuBar::item:selected { background:#2f81f7; color:#ffffff; border-radius:4px; }
            QMenu { background:#161b22; color:#e6edf3; border:1px solid #30363d; }
            QMenu::item:selected { background:#2f81f7; color:#ffffff; }
        """)
        ma = mb.addMenu("Archivo")
        ma.addAction("Guardar Todo", self._save_all)
        ma.addSeparator()
        ma.addAction("Salir", self.close)
        mh = mb.addMenu("Herramientas")
        mh.addAction("Verificar Navegadores", self.tab_nav.check_browsers)
        mh.addAction("Abrir carpeta del programa", lambda: self._abrir_carpeta(_DATA_DIR))
        mh.addAction("Abrir Descargas", lambda: self._abrir_carpeta(_DATA_DIR / "downloads"))
        mh.addSeparator()
        mh.addAction("Buscar Actualizaciones", self._check_updates)
        my = mb.addMenu("Ayuda")
        my.addAction("Descargar Plantilla Excel", self._descargar_plantilla)
        my.addSeparator()
        my.addAction("Acerca de", self._about)

    def _abrir_carpeta(self, ruta):
        try:
            ruta = Path(ruta)
            ruta.mkdir(parents=True, exist_ok=True)
            os.startfile(str(ruta))
        except Exception as ex:
            QMessageBox.warning(self, "Abrir carpeta", f"No se pudo abrir:\n{ruta}\n\n{ex}")

    def _save_all(self):
        self.tab_cred.save_main_cred()
        self.tab_cred._save_hc()
        self.tab_nav.save_config()
        self.statusBar().showMessage("Todo guardado")

    def _descargar_plantilla(self):
        """Genera y guarda la plantilla Excel oficial para Descarga Masiva.
        Incluye la columna NUMERO DE INGRESO y ejemplos de RANGO FECHAS.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception as ex:
            QMessageBox.critical(self, "Falta openpyxl",
                f"No se pudo cargar openpyxl:\n{ex}")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar plantilla Excel",
            str(Path.home() / "Downloads" / "PLANTILLA_BOT.xlsx"),
            "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PACIENTES"
            headers = ["CC", "SERVICIO", "ESTRATEGIA", "FECHA INICIO", "FECHA FIN",
                       "NUMERO DE FACTURA", "NUMERO DE INGRESO"]
            ws.append(headers)
            ejemplos = [
                ["1151937573", "MEDICINA GENERAL", "RECIENTE",     "01/01/2026", "07/05/2026", "", ""],
                ["1151937573", "MEDICINA GENERAL", "RANGO FECHAS", "01/01/2026", "07/05/2026", "", ""],
                ["1151937573", "MEDICINA GENERAL", "RECIENTE",     "01/01/2026", "07/05/2026", "", "8939140"],
                ["52123456",   "PSIQUIATRIA",      "EVOLUCION",    "01/03/2026", "30/04/2026", "", ""],
                ["80123456",   "PSICOLOGIA CONTROL SM", "ANTIGUA", "01/01/2026", "07/05/2026", "", ""],
            ]
            for f in ejemplos:
                ws.append(f)
            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill("solid", fgColor="1F4E78")
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=1, column=col)
                c.font = hf; c.fill = hfill
                c.alignment = Alignment(horizontal="center")
            for i, w in enumerate([14, 24, 16, 14, 14, 22, 22], start=1):
                ws.column_dimensions[chr(64 + i)].width = w

            ws2 = wb.create_sheet("INSTRUCCIONES")
            inst = [
                ["Columna", "Obligatoria", "Descripcion"],
                ["CC", "SI", "Cedula del paciente."],
                ["SERVICIO", "Recomendado", "Nombre EXACTO del servicio en INPEC."],
                ["ESTRATEGIA", "No", "RECIENTE | ANTIGUA | RANGO FECHAS | EVOLUCION | VALORACION | PRIMERA VEZ."],
                ["FECHA INICIO", "No*", "DD/MM/YYYY. Si vacia se usa la fecha global del UI."],
                ["FECHA FIN", "No*", "DD/MM/YYYY. Si vacia se usa la fecha global del UI."],
                ["NUMERO DE FACTURA", "No", "Si se llena, se anexa al nombre del PDF."],
                ["NUMERO DE INGRESO", "No", "Si se llena: descarga SOLO la HC con ese ingreso."],
                ["", "", ""],
                ["MODO RANGO FECHAS", "", "Descarga TODAS las HC ATENDIDAS en el rango. Nombre: CC_SERVICIO_INGRESO.pdf"],
                ["ESTADOS DESCARTADOS", "", "NO ASISTIDA, CANCELADA, POR ATENDER (no se descargan)."],
            ]
            for fila in inst:
                ws2.append(fila)
            for col in range(1, 4):
                c = ws2.cell(row=1, column=col)
                c.font = hf; c.fill = hfill
            ws2.column_dimensions["A"].width = 22
            ws2.column_dimensions["B"].width = 14
            ws2.column_dimensions["C"].width = 95

            wb.save(path)
            self.statusBar().showMessage(f"Plantilla guardada: {path}")
            r = QMessageBox.question(self, "Plantilla creada",
                f"Plantilla generada en:\n{path}\n\n¿Abrirla ahora?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if r == QMessageBox.StandardButton.Yes:
                try:
                    os.startfile(path)
                except Exception:
                    pass
        except Exception as ex:
            QMessageBox.critical(self, "Error",
                f"No se pudo crear la plantilla:\n{ex}")

    def _check_updates(self):
        self.statusBar().showMessage("Verificando actualizaciones...")
        UPDATE_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"

        def do_check():
            try:
                import urllib.request
                import json as _json
                import tempfile

                with urllib.request.urlopen(UPDATE_URL, timeout=8) as resp:
                    data = _json.loads(resp.read().decode())

                latest_tag = data.get("tag_name", "")
                latest = latest_tag.lstrip("v")

                if not latest or latest == VERSION:
                    QTimer.singleShot(0, lambda: self.statusBar().showMessage(
                        f"Ya tienes la versión más reciente (v{VERSION})"))
                    return

                # Buscar el asset .exe directo (onefile build)
                assets = data.get("assets", [])
                exe_asset = next(
                    (a for a in assets if a.get("name", "").lower().endswith(".exe")), None
                )
                download_url = exe_asset["browser_download_url"] if exe_asset else None
                total_size   = (exe_asset.get("size") or 0) if exe_asset else 0

                def ask():
                    self.statusBar().showMessage(f"Nueva versión disponible: v{latest}")
                    size_mb = f"{total_size / 1_048_576:.1f} MB" if total_size else ""
                    if not download_url:
                        QMessageBox.information(self, "Actualización disponible",
                            f"Nueva versión: v{latest}\nActual: v{VERSION}\n\n"
                            "No se encontró el .exe para descarga directa.\n"
                            f"Descárgala en:\nhttps://github.com/{GITHUB_REPO}/releases")
                        return

                    if not getattr(sys, "frozen", False):
                        QMessageBox.information(self, "Actualización disponible",
                            f"Nueva versión: v{latest}  (actual: v{VERSION})\n\n"
                            "Estás ejecutando desde código fuente: actualiza con git.\n"
                            f"Release: https://github.com/{GITHUB_REPO}/releases")
                        return

                    r = QMessageBox.question(self, "Actualización disponible",
                        f"Nueva versión: v{latest}  (actual: v{VERSION})\n"
                        f"Tamaño: {size_mb}\n\n"
                        "¿Actualizar ahora?\n\n"
                        "El programa se cerrará, se actualizará y se reabrirá automáticamente.\n"
                        "Tu configuración (config.json) no se toca.",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                    if r != QMessageBox.StandardButton.Yes:
                        return

                    def download_and_replace():
                        try:
                            # 1. Descargar nuevo .exe a archivo temporal
                            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".exe", prefix="BOT_new_")
                            os.close(tmp_fd)
                            received = 0

                            with urllib.request.urlopen(download_url, timeout=120) as dl, \
                                 open(tmp_path, "wb") as f:
                                while True:
                                    chunk = dl.read(65536)
                                    if not chunk:
                                        break
                                    f.write(chunk)
                                    received += len(chunk)
                                    if total_size:
                                        pct = int(received * 100 / total_size)
                                        QTimer.singleShot(0, lambda p=pct:
                                            self.statusBar().showMessage(
                                                f"Descargando actualización... {p}%"))

                            # 2. Obtener ruta del .exe actual
                            current_exe = Path(sys.executable if getattr(sys, "frozen", False)
                                               else __file__).resolve()

                            # 3. Crear script batch que espera cierre, reemplaza y relanza
                            bat_fd, bat_path = tempfile.mkstemp(suffix=".bat", prefix="bot_upd_")
                            os.close(bat_fd)
                            bat_content = (
                                "@echo off\n"
                                "timeout /t 2 /nobreak >nul\n"
                                f'move /Y "{tmp_path}" "{current_exe}"\n'
                                f'start "" "{current_exe}"\n'
                                f'del "%~f0"\n'
                            )
                            with open(bat_path, "w") as bf:
                                bf.write(bat_content)

                            # 4. Lanzar script y cerrar app
                            subprocess.Popen(
                                ["cmd", "/c", bat_path],
                                creationflags=subprocess.CREATE_NO_WINDOW,
                            )
                            QTimer.singleShot(0, lambda: QApplication.instance().quit())

                        except Exception as ex:
                            _msg_ex = str(ex)
                            QTimer.singleShot(0, lambda: (
                                self.statusBar().showMessage(f"Error al actualizar: {_msg_ex}"),
                                QMessageBox.critical(self, "Error de actualización",
                                    f"No se pudo completar la actualización:\n{_msg_ex}"),
                            ))

                    threading.Thread(target=download_and_replace, daemon=True).start()

                QTimer.singleShot(0, ask)

            except Exception as ex:
                _msg_ex_verif = str(ex)
                QTimer.singleShot(0, lambda: self.statusBar().showMessage(
                    f"Error al verificar: {_msg_ex_verif}"))

        threading.Thread(target=do_check, daemon=True).start()

    def _about(self):
        QMessageBox.about(
            self, "Acerca de BOT",
            f"<b>BOT</b> v{VERSION}<br><br>"
            "Automatizacion de Agendas y Gestion de HC.<br>"
            "Python 3.9+  |  PyQt6  |  Selenium<br><br>"
            "Licencia: MIT"
        )

    def closeEvent(self, event):
        r = QMessageBox.question(
            self, "Salir", "Deseas cerrar BOT?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        event.accept() if r == QMessageBox.StandardButton.Yes else event.ignore()


# ===========================================================================
# Entry point
# ===========================================================================
def _escribir_error_txt(titulo: str, detalle: str) -> str:
    """Escribe el UNICO archivo que el bot crea ante un error: BOT_error.txt,
    junto al ejecutable, con el error EXACTO. No crea carpetas ni nada mas.
    Devuelve la ruta escrita (o '' si no se pudo)."""
    try:
        import datetime as _dt
        ruta = _DATA_DIR / "BOT_error.txt"
        sello = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(ruta, "w", encoding="utf-8") as f:
            f.write(f"BOT v{VERSION} - ERROR\n")
            f.write(f"Fecha: {sello}\n")
            f.write(f"{titulo}\n")
            f.write("=" * 60 + "\n")
            f.write(detalle.rstrip() + "\n")
        return str(ruta)
    except Exception:
        return ""


def _instalar_manejador_errores():
    """Captura cualquier excepcion NO controlada y la vuelca a BOT_error.txt
    (en vez de crashear en silencio en otro PC)."""
    import traceback

    def _hook(exc_type, exc, tb):
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc, tb)
            return
        detalle = "".join(traceback.format_exception(exc_type, exc, tb))
        ruta = _escribir_error_txt("Excepcion no controlada", detalle)
        try:
            from PyQt6.QtWidgets import QApplication as _QApp, QMessageBox as _QMB
            if _QApp.instance() is not None:
                _QMB.critical(None, "Error", f"Ocurrio un error.\nDetalle guardado en:\n{ruta or 'BOT_error.txt'}")
        except Exception:
            pass

    sys.excepthook = _hook


def main():
    _instalar_manejador_errores()
    try:
        app = QApplication(sys.argv)
        app.setApplicationName("BOT")
        app.setApplicationVersion(VERSION)
        win = MainWindow()
        sys.exit(app.exec())
    except SystemExit:
        raise
    except BaseException as exc:
        import traceback
        ruta = _escribir_error_txt(
            "Fallo al iniciar la aplicacion",
            "".join(traceback.format_exception(type(exc), exc, exc.__traceback__)),
        )
        # Intentar avisar visualmente; si ni eso se puede, queda el txt.
        try:
            from PyQt6.QtWidgets import QApplication as _QApp, QMessageBox as _QMB
            _app = _QApp.instance() or _QApp(sys.argv)
            _QMB.critical(None, "Error al iniciar",
                          f"No se pudo iniciar BOT.\nError guardado en:\n{ruta or 'BOT_error.txt'}")
        except Exception:
            pass
        raise


if __name__ == "__main__":
    main()
