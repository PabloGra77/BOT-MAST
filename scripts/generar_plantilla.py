import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT = r"c:\Users\LUIs CORDOBA\Downloads\PLANTILLA_BOT360_v10.0.22.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PACIENTES"

headers = ["CC", "SERVICIO", "ESTRATEGIA", "FECHA INICIO", "FECHA FIN",
           "NUMERO DE FACTURA", "NUMERO DE INGRESO"]
ws.append(headers)

filas = [
    ["1151937573", "MEDICINA GENERAL", "RECIENTE",     "01/01/2026", "07/05/2026", "", ""],
    ["1151937573", "MEDICINA GENERAL", "RANGO FECHAS", "01/01/2026", "07/05/2026", "", ""],
    ["1151937573", "MEDICINA GENERAL", "RECIENTE",     "01/01/2026", "07/05/2026", "", "8939140"],
    ["52123456",   "PSIQUIATRIA",      "EVOLUCION",    "01/03/2026", "30/04/2026", "", ""],
    ["80123456",   "PSICOLOGIA CONTROL SM", "ANTIGUA", "01/01/2026", "07/05/2026", "", ""],
]
for f in filas:
    ws.append(f)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
for col_idx in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

anchos = [14, 24, 16, 14, 14, 22, 22]
for i, w in enumerate(anchos, start=1):
    ws.column_dimensions[chr(64 + i)].width = w

ws2 = wb.create_sheet("INSTRUCCIONES")
inst = [
    ["Columna", "Obligatoria", "Descripcion"],
    ["CC", "SI", "Cedula del paciente."],
    ["SERVICIO", "Recomendado", "Nombre EXACTO del servicio en INPEC360 (ej: MEDICINA GENERAL, PSIQUIATRIA)."],
    ["ESTRATEGIA", "No", "RECIENTE (default) | ANTIGUA | RANGO FECHAS | EVOLUCION | VALORACION | PRIMERA VEZ."],
    ["FECHA INICIO", "No*", "DD/MM/YYYY. Si vacia se usa la fecha global del UI."],
    ["FECHA FIN", "No*", "DD/MM/YYYY. Si vacia se usa la fecha global del UI."],
    ["NUMERO DE FACTURA", "No", "Si se llena, se anexa al nombre del PDF."],
    ["NUMERO DE INGRESO", "No", "Si se llena: descarga SOLO la HC con ese # de ingreso. Si vacia: comportamiento normal."],
    ["", "", ""],
    ["MODO RANGO FECHAS", "", "Descarga TODAS las HC validas (estado ATENDIDO) en el rango. Cada PDF se nombra CC_SERVICIO_INGRESO.pdf."],
    ["ESTADOS DESCARTADOS", "", "NO ASISTIDA, CANCELADA, POR ATENDER (no se descargan)."],
]
for fila in inst:
    ws2.append(fila)
for col_idx in range(1, 4):
    c = ws2.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 95

wb.save(OUT)
print("OK ->", OUT)
